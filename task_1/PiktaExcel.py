import os.path
from datetime import datetime

import openpyxl
from openpyxl import Workbook


class PiktaExcel:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.path_is_changed = False

    def add_worksheet(self, sheet_name: str, data: list):
        excel_file = openpyxl.load_workbook(self.excel_file_path)
        excel_sheet = excel_file.create_sheet(title=sheet_name, index=0)
        for row in data:
            excel_sheet.append(row)
        excel_file.save(self.excel_file_path)

    def create_excel(self):
        path = self.excel_file_path
        if self.__check_path_is_exist(path):
            path = self.__create_new_path(path)
        excel_file = Workbook()
        excel_file.save(path)

    def __change_path(self, new_path: str):
        self.path_is_changed = True
        self.excel_file_path = new_path

    def __check_path_is_exist(self, path):
        return os.path.isfile(path)

    def __create_new_path(self, path: str) -> str:
        now = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
        path_split = path.split(".")
        new_path = f"{''.join(path_split[:-1])}_{now}.{path_split[-1]}"
        self.__change_path(new_path)
        return new_path

